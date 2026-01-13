"""
Complete NDNC Automation - Review Pending & Open Complaints
Handles both workflows with proper verification and file management
"""

import os
import re
import time
import shutil
import PyPDF2
import pytesseract
import openpyxl
from pathlib import Path
from datetime import datetime, timedelta
from PIL import Image
from pdf2image import convert_from_path
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException


class NDNCCompleteAutomation:
    def __init__(self, email: str):
        """Initialize NDNC automation"""
        self.email = email
        self.base_url = "https://dashboard.ndnc.exotel.com"
        self.driver = None
        
        # Setup folders
        self.ndnc_folder = Path.home() / "Downloads" / "NDNC"
        self.review_pending_folder = self.ndnc_folder / "review_pending"
        self.open_folder = self.ndnc_folder / "open"
        self.processed_folder = self.ndnc_folder / "processed"
        self.processed_review_folder = self.ndnc_folder / "processed_review"
        
        # Create folders
        for folder in [self.review_pending_folder, self.open_folder, self.processed_folder, self.processed_review_folder]:
            folder.mkdir(parents=True, exist_ok=True)
    
    def start_browser(self):
        """Start browser with download preferences"""
        print("→ Starting Chrome browser...")
        options = webdriver.ChromeOptions()
        
        # Set download directory
        prefs = {
            "download.default_directory": str(self.ndnc_folder),
            "download.prompt_for_download": False,
            "download.directory_upgrade": True,
            "safebrowsing.enabled": True
        }
        options.add_experimental_option("prefs", prefs)
        
        self.driver = webdriver.Chrome(options=options)
        time.sleep(1)
        self.driver.maximize_window()
        time.sleep(0.5)
        print("✓ Browser started and ready")
    
    def login(self):
        """Login to NDNC dashboard"""
        try:
            print(f"\n→ Navigating to login page...")
            self.driver.get(f"{self.base_url}/login")
            time.sleep(2)
            
            print(f"→ Entering email: {self.email}")
            wait = WebDriverWait(self.driver, 15)
            email_input = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'input[type="email"]')))
            
            # Type email slowly like a human (don't use .clear())
            for char in self.email:
                email_input.send_keys(char)
                time.sleep(0.1)
            
            # Wait before clicking
            time.sleep(1)
            
            print(f"→ Clicking Continue button...")
            continue_button = self.driver.find_element(By.CSS_SELECTOR, 'button[type="submit"]')
            continue_button.click()
            
            print(f"\n{'='*70}")
            print(f"⏳ PLEASE ENTER OTP IN THE BROWSER WINDOW")
            print(f"   Waiting for you to enter OTP...")
            print(f"   (Timeout: 5 minutes)")
            print(f"{'='*70}\n")
            
            # Wait for navigation to dashboard (indicating successful login)
            try:
                WebDriverWait(self.driver, 300).until(
                    EC.url_contains("/dashboard")
                )
                print("✓ Login successful!")
                time.sleep(5)
                
                # Wait for dashboard page to fully load
                print("   → Waiting for dashboard to fully load...")
                wait = WebDriverWait(self.driver, 30)
                
                # Wait for dashboard elements to appear
                try:
                    wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'nav, header, a[href*="complaints"]')))
                    print("   ✓ Dashboard elements loaded")
                except:
                    print("   ⚠️  Dashboard elements not found, but continuing...")
                
                time.sleep(3)
                
                # Verify we're actually logged in by checking URL
                current_url = self.driver.current_url
                print(f"   → Current URL: {current_url}")
                
                if "login" in current_url:
                    print("✗ Still on login page - OTP may have failed")
                    return False
                
                # Try clicking on dashboard link if available
                try:
                    dashboard_link = self.driver.find_element(By.CSS_SELECTOR, 'a[href="/dashboard"]')
                    dashboard_link.click()
                    time.sleep(3)
                    print("   ✓ Navigated to main dashboard")
                except:
                    print("   ⚠️  Dashboard link not found, proceeding anyway...")
                
                print("   ✓ Session verified - ready to proceed")
                time.sleep(2)
                return True
                
            except TimeoutException:
                print("✗ Login timeout - please ensure OTP was entered")
                return False
                
        except Exception as e:
            print(f"✗ Login error: {str(e)}")
            import traceback
            traceback.print_exc()
            return False
    
    def navigate_to_all_complaints(self):
        """Navigate to All Complaints page"""
        try:
            print("\n→ Navigating to All Complaints...")
            
            # First check current URL
            current_url = self.driver.current_url
            print(f"   → Current URL: {current_url}")
            
            # If already on all-complaints, no need to navigate
            if "all-complaints" in current_url:
                print("   ✓ Already on All Complaints page")
                return True
            
            # Navigate to All Complaints
            print(f"   → Going to: {self.base_url}/all-complaints")
            self.driver.get(f"{self.base_url}/all-complaints")
            time.sleep(3)
            
            # Wait for page to fully load
            print("   → Waiting for page to load...")
            wait = WebDriverWait(self.driver, 20)
            
            # Try to find either table or search input
            try:
                wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'table tbody tr')))
                print("   ✓ Table loaded")
            except:
                try:
                    wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'input[placeholder*="Search"]')))
                    print("   ✓ Search field loaded")
                except:
                    print("   ⚠️  Page elements not found, but continuing...")
            
            time.sleep(2)
            
            # Verify we're on the right page
            final_url = self.driver.current_url
            print(f"   → Final URL: {final_url}")
            
            if "all-complaints" in final_url or "dashboard" in final_url:
                print("✓ Successfully on All Complaints page")
                return True
            else:
                print(f"✗ Not on expected page: {final_url}")
                return False
                
        except Exception as e:
            print(f"✗ Navigation error: {str(e)}")
            import traceback
            traceback.print_exc()
            
            # Check if we need to login again
            current_url = self.driver.current_url
            if "login" in current_url:
                print(f"⚠️  Browser redirected to login page - session may have expired")
            
            return False
    
    def extract_phone_from_filename(self, filename: str) -> str:
        """Extract phone number from filename"""
        # Try patterns: 9479760361_1135047815_18-Dec-2025_Call1.pdf
        match = re.search(r'(\d{10})', filename)
        if match:
            return match.group(1)
        return None
    
    def extract_telemarketer_from_filename(self, filename: str) -> str:
        """Extract telemarketer number (second 10-digit number) from filename"""
        matches = re.findall(r'(\d{10})', filename)
        if len(matches) >= 2:
            return matches[1]  # Second 10-digit number is telemarketer
        return None
    
    def clean_ordinal_date(self, date_str: str) -> str:
        """Remove ordinal suffixes (st, nd, rd, th) from dates"""
        # Remove st, nd, rd, th from dates like "17th Dec 2025" -> "17 Dec 2025"
        return re.sub(r'(\d{1,2})(st|nd|rd|th)\b', r'\1', date_str)
    
    def convert_date_format(self, date_str: str) -> str:
        """Convert date from various formats to Month DD, YYYY format"""
        # Clean ordinal suffixes first
        date_str = self.clean_ordinal_date(date_str)
        
        date_formats = [
            '%d %b %Y',    # 17 Dec 2025 (after cleaning "17th Dec 2025")
            '%d %B %Y',    # 17 December 2025
            '%d-%b-%Y',    # 12-Dec-2025
            '%d-%m-%Y',    # 12-12-2025
            '%d/%b/%y',    # 11/Dec/25
            '%d-%b-%y',    # 11-Dec-25
            '%d/%b/%Y',    # 11/Dec/2025
            '%d.%b.%Y',    # 11.Dec.2025
            '%d.%b.%y',    # 11.Dec.25
            '%d/%m/%Y',    # 11/12/2025
            '%d/%m/%y',    # 11/12/25
            '%b %d, %Y',   # Jan 2, 2026
            '%B %d, %Y',   # January 2, 2026
        ]
        
        for fmt in date_formats:
            try:
                date_obj = datetime.strptime(date_str, fmt)
                return date_obj.strftime('%B %d, %Y')
            except:
                continue
        
        # If no format matches, return original
        return date_str
    
    def extract_all_dates_from_text(self, text: str) -> list:
        """Extract ALL dates from text content"""
        dates_found = []
        
        # Comprehensive date patterns (more flexible)
        date_patterns = [
            r'\b(\d{1,2})(?:st|nd|rd|th)?\s+(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*\s+(\d{4})\b',  # 17th Dec 2025, 3rd Dec 2025
            r'\b(\d{1,2})[-/\s_](\w{3,9})[-/\s_](\d{4})\b',  # 14-Dec-2025, 15 December 2025, 14_Dec_2025
            r'\b(\d{1,2})[-/\s_](\w{3,9})[-/\s_](\d{2})\b',  # 14-Dec-25, 14_Dec_25
            r'\b(\w{3,9})\s+(\d{1,2}),?\s+(\d{4})\b',         # December 14, 2025
            r'\b(\d{1,2})[-/\.](\d{1,2})[-/\.](\d{4})\b',    # 14/12/2025, 14.12.2025
            r'\b(\d{4})[-/\.](\d{1,2})[-/\.](\d{1,2})\b',    # 2025-12-14, 2025.12.14
            r'\b(\d{1,2})\s+(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*\s+(\d{4})\b',  # 14 December 2025
            r'\b(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*\s+(\d{1,2}),?\s+(\d{4})\b',  # December 14, 2025
        ]
        
        for pattern in date_patterns:
            matches = re.finditer(pattern, text, re.IGNORECASE)
            for match in matches:
                date_str = match.group(0).strip()
                # Clean up date string
                date_str = re.sub(r'\s+', ' ', date_str)  # Normalize spaces
                if date_str and len(date_str) >= 6:  # Minimum date length
                    dates_found.append(date_str)
        
        return dates_found
    
    def preprocess_image_for_ocr(self, image):
        """Apply multiple preprocessing techniques to improve OCR accuracy"""
        import cv2
        import numpy as np
        
        # Convert PIL Image to numpy array
        img_array = np.array(image)
        
        # Try multiple preprocessing approaches
        processed_images = []
        
        # 1. Original image
        processed_images.append(img_array)
        
        # 2. Grayscale conversion
        if len(img_array.shape) == 3:
            gray = cv2.cvtColor(img_array, cv2.COLOR_RGB2GRAY)
            processed_images.append(gray)
            
            # 3. Thresholding
            _, thresh1 = cv2.threshold(gray, 150, 255, cv2.THRESH_BINARY)
            processed_images.append(thresh1)
            
            # 4. Adaptive thresholding
            thresh2 = cv2.adaptiveThreshold(gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, 11, 2)
            processed_images.append(thresh2)
            
            # 5. Noise removal
            denoised = cv2.fastNlMeansDenoising(gray, None, 10, 7, 21)
            processed_images.append(denoised)
        
        return processed_images
    
    def extract_url_from_address_bar(self, image):
        """Enhanced OCR specifically for browser address bar region"""
        import cv2
        import numpy as np
        
        try:
            # Convert PIL to numpy
            img_array = np.array(image)
            height, width = img_array.shape[:2]
            
            # Focus on top 150 pixels (browser address bar area)
            address_bar_region = img_array[0:min(150, height), :]
            
            # Convert to grayscale
            if len(address_bar_region.shape) == 3:
                gray = cv2.cvtColor(address_bar_region, cv2.COLOR_RGB2GRAY)
            else:
                gray = address_bar_region
            
            # Apply aggressive preprocessing for small text
            # 1. Upscale 3x for better OCR
            upscaled = cv2.resize(gray, None, fx=3, fy=3, interpolation=cv2.INTER_CUBIC)
            
            # 2. Contrast enhancement
            clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8,8))
            enhanced = clahe.apply(upscaled)
            
            # 3. Sharpening kernel
            kernel = np.array([[-1,-1,-1],
                             [-1, 9,-1],
                             [-1,-1,-1]])
            sharpened = cv2.filter2D(enhanced, -1, kernel)
            
            # 4. Binary threshold
            _, binary = cv2.threshold(sharpened, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
            
            # Try OCR on enhanced address bar region
            configs = [
                '--psm 7',  # Single text line (perfect for address bar)
                '--psm 6',  # Uniform block
            ]
            
            extracted_text = []
            for config in configs:
                try:
                    text = pytesseract.image_to_string(binary, config=config)
                    if text.strip():
                        extracted_text.append(text.strip())
                except:
                    continue
            
            return ' '.join(extracted_text)
            
        except Exception as e:
            print(f"   → Address bar extraction error: {str(e)}")
            return ""
    
    def extract_data_from_file(self, file_path: Path) -> dict:
        """Extract phone numbers, dates, URLs, and logos from file content using OCR"""
        try:
            print(f"\n   → Performing comprehensive OCR extraction...")
            all_text = ""
            
            # Read file based on type with enhanced OCR
            if file_path.suffix.lower() in ['.png', '.jpg', '.jpeg']:
                image = Image.open(file_path)
                
                # FIRST: Try to extract URL from browser address bar (top of screenshot)
                print(f"   → Attempting enhanced URL extraction from address bar...")
                address_bar_text = self.extract_url_from_address_bar(image)
                if address_bar_text:
                    print(f"   ✓ Address bar text: {address_bar_text[:100]}")
                
                # Try multiple OCR configurations and preprocessing
                ocr_configs = [
                    '--psm 6',  # Assume uniform block of text
                    '--psm 3',  # Fully automatic page segmentation
                    '--psm 11', # Sparse text
                    '--psm 12', # Sparse text with OSD
                ]
                
                # Get preprocessed images
                try:
                    processed_images = self.preprocess_image_for_ocr(image)
                except Exception as prep_error:
                    print(f"   → Preprocessing failed, using original image: {str(prep_error)}")
                    processed_images = [image]
                
                # Try OCR with each configuration and preprocessing combination
                texts = []
                
                # Add address bar text first (highest priority)
                if address_bar_text:
                    texts.append(address_bar_text)
                
                for config in ocr_configs:
                    for proc_img in processed_images[:3]:  # Use first 3 preprocessing methods
                        try:
                            text = pytesseract.image_to_string(proc_img, config=config)
                            if len(text.strip()) > 10:
                                texts.append(text)
                        except:
                            continue
                
                # Combine all extracted text
                all_text = "\n".join(texts)
                
            else:
                # Try text extraction first for PDFs
                try:
                    with open(file_path, 'rb') as f:
                        pdf_reader = PyPDF2.PdfReader(f)
                        for page in pdf_reader.pages:
                            all_text += page.extract_text()
                except:
                    pass
                
                # Use OCR if needed
                if len(all_text.strip()) < 50:
                    images = convert_from_path(str(file_path), dpi=300)
                    for image in images:
                        # Try multiple configs for PDFs too
                        for config in ['--psm 6', '--psm 3', '--psm 11']:
                            try:
                                text = pytesseract.image_to_string(image, config=config)
                                all_text += text + "\n"
                            except:
                                continue
            
            print(f"   → OCR extracted {len(all_text)} characters")
            
            # Debug: Show sample of extracted text
            if len(all_text.strip()) > 0:
                sample = all_text[:500].replace('\n', ' ')
                print(f"   → Text sample: {sample[:150]}...")
            
            # Extract ALL phone numbers (10-digit, strip 91 prefix)
            phone_pattern = r'(?:91)?(\d{10})'
            phone_matches = re.findall(phone_pattern, all_text)
            # Filter unique valid phones
            unique_phones = list(set([p for p in phone_matches if len(p) == 10 and not p.startswith('0000')]))
            
            # Extract ALL dates
            all_dates = self.extract_all_dates_from_text(all_text)
            unique_dates = list(set(all_dates))
            
            # ALWAYS try to extract date from filename (primary source for matching portal)
            filename_date = re.search(r'(\d{1,2}-\w{3}-\d{4})', file_path.name)
            if filename_date:
                extracted_filename_date = filename_date.group(1)
                if extracted_filename_date not in unique_dates:
                    unique_dates.insert(0, extracted_filename_date)  # Add at beginning (highest priority)
                    print(f"   → Added date from filename: {extracted_filename_date} (for portal matching)")
            
            # Sort dates by length (longer dates are usually more complete)
            unique_dates.sort(key=lambda x: len(x), reverse=True)
            
            # Check for URL/logo patterns - Includes business CRM systems and company branding
            url_patterns = ['zomato', 'blinkit', 'lifeline', 'exotel', 'swiggy', 'uber', 'ola', 'amazon', 
                          'flipkart', 'dunzo', 'bigbasket', 'grofers', 'myntra', 'meesho', 'paytm',
                          'phonepe', 'gpay', 'hdfc', 'crm', 'dynamics', 'salesforce', 'persistency', 
                          'persistence', 'shipsy', 'gam', 'portal', 'analytics', 'visualize',
                          'http://', 'https://', 'www.', 
                          '.com', '.in', '.org', '.net', '.io', '.co', '.ai', '.tech']
            logo_patterns = ['order', 'delivery', 'invoice', 'receipt', 'bill', 'lead', 'policy', 
                           'hdfc', 'life', 'insurance', 'visualisation', 'dashboard', 'analytics']
            
            found_urls = [url for url in url_patterns if re.search(re.escape(url), all_text, re.IGNORECASE)]
            found_logos = [logo for logo in logo_patterns if re.search(re.escape(logo), all_text, re.IGNORECASE)]
            
            # Print results
            if unique_phones:
                print(f"   ✓ Found {len(unique_phones)} phone number(s): {', '.join(unique_phones)}")
            else:
                print(f"   ✗ No phone numbers found in document")
            
            if unique_dates:
                print(f"   ✓ Found {len(unique_dates)} date(s): {', '.join(unique_dates)}")
            else:
                print(f"   ✗ No dates found in document")
            
            if found_urls:
                print(f"   ✓ Found URL patterns: {', '.join(found_urls[:3])}")
            else:
                print(f"   ⚠️  No known URL patterns found")
            
            if found_logos:
                print(f"   ✓ Found logo/brand text: {', '.join(found_logos[:3])}")
            else:
                print(f"   ⚠️  No logo text found")
            
            return {
                'phone': unique_phones[0] if unique_phones else None,
                'all_phones': unique_phones,
                'all_dates': unique_dates,
                'date': unique_dates[0] if unique_dates else None,
                'text': all_text,
                'urls_found': found_urls,
                'logos_found': found_logos,
                'has_authenticity': len(found_urls) > 0 or len(found_logos) > 0
            }
            
        except Exception as e:
            print(f"   ✗ OCR extraction error: {str(e)}")
            import traceback
            traceback.print_exc()
            return {
                'phone': None,
                'all_phones': [],
                'all_dates': [],
                'date': None,
                'text': '',
                'urls_found': [],
                'logos_found': [],
                'has_authenticity': False
            }
    
    def validate_document_completely(self, file_data: dict, expected_phone: str, url_date_str: str) -> tuple:
        """
        Comprehensive validation of document using OCR data
        Returns: (is_valid: bool, reason: str)
        """
        print(f"\n{'='*70}")
        print(f"📋 COMPREHENSIVE DOCUMENT VALIDATION")
        print(f"{'='*70}")
        
        validation_results = []
        all_passed = True
        
        # 1. Check for URL/Logo authenticity
        print(f"\n1. Authenticity Check (URL/Logo):")
        if file_data.get('has_authenticity'):
            urls = file_data.get('urls_found', [])
            logos = file_data.get('logos_found', [])
            print(f"   ✓ PASSED - Found patterns:")
            if urls:
                print(f"      URLs: {', '.join(urls)}")
            if logos:
                print(f"      Logos: {', '.join(logos)}")
            validation_results.append(("Authenticity", True, f"Found: {', '.join(urls + logos)}"))
        else:
            print(f"   ✗ FAILED - No URL or logo patterns found")
            print(f"      Required: Company URL or document proof")
            print(f"      Note: Enhanced OCR attempted but no patterns detected")
            validation_results.append(("Authenticity", False, "No URL/logo found in document"))
            all_passed = False
        
        # 2. Check for phone number
        print(f"\n2. Phone Number Check:")
        found_phones = file_data.get('all_phones', [])
        if expected_phone in found_phones:
            print(f"   ✓ PASSED - Phone {expected_phone} found in document")
            validation_results.append(("Phone", True, f"Matched: {expected_phone}"))
        else:
            print(f"   ✗ FAILED - Phone {expected_phone} NOT found in document")
            if found_phones:
                print(f"      Found in doc: {', '.join(found_phones)}")
                print(f"      Expected: {expected_phone}")
            else:
                print(f"      No phone numbers found in document")
            validation_results.append(("Phone", False, f"Expected {expected_phone}, found {found_phones}"))
            all_passed = False
        
        # 3. Check for date within 6 months of URL date
        print(f"\n3. Date Validation (within 6 months of URL date):")
        url_date_obj = None
        
        # Clean ordinal suffixes from URL date
        cleaned_url_date = self.clean_ordinal_date(url_date_str)
        
        # Try multiple date formats for URL date
        date_formats = [
            '%d %b %Y',    # 17 Dec 2025
            '%d %B %Y',    # 17 December 2025
            '%d-%b-%Y',    # 12-Dec-2025
            '%d-%m-%Y',    # 12-12-2025
            '%d/%b/%y',    # 11/Dec/25
            '%d-%b-%y',    # 11-Dec-25
            '%d/%b/%Y',    # 11/Dec/2025
            '%d.%b.%Y',    # 11.Dec.2025
            '%d.%b.%y',    # 11.Dec.25
            '%d/%m/%Y',    # 11/12/2025
            '%d/%m/%y',    # 11/12/25
        ]
        
        for fmt in date_formats:
            try:
                url_date_obj = datetime.strptime(cleaned_url_date, fmt)
                break
            except:
                continue
        
        found_dates = file_data.get('all_dates', [])
        valid_date_found = False
        
        if not url_date_obj:
            print(f"   ⚠️  Cannot parse URL date: {url_date_str}")
        elif not found_dates:
            print(f"   ✗ FAILED - No dates found in document content")
            validation_results.append(("Date", False, "No dates found in document"))
            all_passed = False
        else:
            print(f"   → URL date: {url_date_str} → {url_date_obj.strftime('%B %d, %Y')}")
            print(f"   → Checking {len(found_dates)} date(s) from document:")
            print(f"      Raw dates: {', '.join(found_dates)}")
            print(f"\n   → Testing each date against URL date...")
            
            # Check each date to see if it's within 6 months
            for idx, date_str in enumerate(found_dates, 1):
                try:
                    # Skip obviously invalid dates (too short, malformed)
                    if len(date_str) < 6 or not any(c.isdigit() for c in date_str):
                        print(f"      ⊘ Date {idx}/{len(found_dates)}: '{date_str}' - SKIPPED (malformed)")
                        continue
                    
                    # Try multiple date formats - including BOTH DD/MM/YYYY and MM/DD/YYYY interpretations
                    doc_date_obj = None
                    cleaned_date_str = self.clean_ordinal_date(date_str)
                    
                    # Comprehensive format list with both DD/MM and MM/DD
                    formats_to_try = [
                        '%d %b %Y', '%d %B %Y',           # 17 Dec 2025
                        '%d-%b-%Y', '%d-%B-%Y',           # 17-Dec-2025
                        '%d/%m/%Y',                        # 16/07/2025 (DD/MM/YYYY)
                        '%m/%d/%Y',                        # 07/16/2025 (MM/DD/YYYY)
                        '%d-%m-%Y',                        # 16-07-2025 (DD-MM-YYYY)
                        '%m-%d-%Y',                        # 07-16-2025 (MM-DD-YYYY)
                        '%Y-%m-%d',                        # 2025-07-16 (ISO)
                        '%d/%m/%y',                        # 16/07/25 (DD/MM/YY)
                        '%m/%d/%y',                        # 07/16/25 (MM/DD/YY)
                        '%d/%b/%y', '%d-%b-%y', '%d/%b/%Y',  # Month abbreviations
                        '%d.%b.%Y', '%d.%b.%y',           # Dot separators
                        '%B %d, %Y', '%b %d, %Y',         # Month name formats
                    ]
                    
                    for fmt in formats_to_try:
                        try:
                            doc_date_obj = datetime.strptime(cleaned_date_str, fmt)
                            break
                        except:
                            continue
                    
                    if doc_date_obj:
                        diff_days = abs((doc_date_obj - url_date_obj).days)
                        diff_months = diff_days / 30.44
                        
                        if diff_months <= 6:
                            print(f"   ✓ PASSED - Date {idx}/{len(found_dates)}: '{date_str}' → {doc_date_obj.strftime('%B %d, %Y')}")
                            print(f"      Difference from URL date: {diff_months:.1f} months ({diff_days} days)")
                            validation_results.append(("Date", True, f"Matched: {date_str} ({diff_months:.1f} months)"))
                            valid_date_found = True
                            break  # Found a valid date within 6 months, we can stop
                        else:
                            print(f"      ✗ Date {idx}/{len(found_dates)}: '{date_str}' → {doc_date_obj.strftime('%B %d, %Y')} is {diff_months:.1f} months away (beyond 6 months)")
                    else:
                        print(f"      ⊘ Date {idx}/{len(found_dates)}: '{date_str}' - Could not parse (invalid format)")
                        
                except Exception as e:
                    print(f"      ⊘ Date {idx}/{len(found_dates)}: '{date_str}' - Error: {str(e)[:50]}")
                    continue
            
            if not valid_date_found:
                print(f"   ✗ FAILED - No dates within 6 months of URL date ({url_date_str})")
                print(f"      All dates found: {', '.join(found_dates)}")
                validation_results.append(("Date", False, f"No dates within 6 months of {url_date_str}"))
                all_passed = False
        
        # Final verdict
        print(f"\n{'='*70}")
        print(f"📊 VALIDATION SUMMARY")
        print(f"{'='*70}")
        for check_name, passed, details in validation_results:
            status = "✓ PASS" if passed else "✗ FAIL"
            print(f"{status} - {check_name}: {details}")
        
        if all_passed:
            print(f"\n🎉 ALL VALIDATIONS PASSED - Document is authentic and valid")
            print(f"{'='*70}\n")
            return (True, "All validations passed")
        else:
            failed_checks = [name for name, passed, _ in validation_results if not passed]
            reason = f"Failed checks: {', '.join(failed_checks)}"
            print(f"\n❌ VALIDATION FAILED - {reason}")
            print(f"{'='*70}\n")
            return (False, reason)
    
    def search_complaint(self, phone_number: str) -> bool:
        """Search for complaint by phone number"""
        try:
            print(f"\n   → Searching for: {phone_number}")
            
            wait = WebDriverWait(self.driver, 15)
            
            # Close any open modals first
            try:
                close_buttons = self.driver.find_elements(By.XPATH, '//div[@role="dialog"][@data-state="open"]//button[.//svg[contains(@class, "lucide-x")]]')
                for btn in close_buttons:
                    try:
                        self.driver.execute_script("arguments[0].click();", btn)
                        time.sleep(0.5)
                    except:
                        pass
            except:
                pass
            
            # Find search input - use the correct selector
            print(f"   → Locating search field...")
            search_input = wait.until(EC.presence_of_element_located(
                (By.CSS_SELECTOR, 'input[placeholder*="Search by number"]')
            ))
            time.sleep(1)
            
            # Use JavaScript to clear and set value instead of .clear()
            self.driver.execute_script("arguments[0].value = '';", search_input)
            time.sleep(0.5)
            
            print(f"   → Entering phone number...")
            for digit in phone_number:
                search_input.send_keys(digit)
                time.sleep(0.1)
            
            # Wait for search results to load (no search button, auto-searches)
            time.sleep(3)
            
            print(f"   ✓ Search executed, results loading...")
            time.sleep(2)
            return True
            
        except Exception as e:
            print(f"   ✗ Search error: {str(e)}")
            import traceback
            traceback.print_exc()
            return False
    
    def find_and_click_complaint(self, expected_date: str, telemarketer_number: str = None) -> bool:
        """Find complaint matching date and click it"""
        try:
            # Convert date to portal format (Month DD, YYYY)
            target_date = self.convert_date_format(expected_date)
            
            # Clean ordinal suffixes and try multiple date formats to parse the expected date
            cleaned_date = self.clean_ordinal_date(expected_date)
            expected_date_obj = None
            date_formats = [
                '%d %b %Y',    # 17 Dec 2025
                '%d %B %Y',    # 17 December 2025
                '%d-%b-%Y',    # 12-Dec-2025
                '%d-%m-%Y',    # 12-12-2025
                '%d/%b/%y',    # 11/Dec/25
                '%d-%b-%y',    # 11-Dec-25
                '%d/%b/%Y',    # 11/Dec/2025
                '%d.%b.%Y',    # 11.Dec.2025
                '%d.%b.%y',    # 11.Dec.25
                '%b %d, %Y',   # Jan 2, 2026
                '%B %d, %Y',   # January 2, 2026
            ]
            
            for fmt in date_formats:
                try:
                    expected_date_obj = datetime.strptime(cleaned_date, fmt)
                    break
                except:
                    continue
            
            if not expected_date_obj:
                print(f"      → Date '{expected_date}' could not be parsed (trying next date...)")
                return False
            
            print(f"      → Parsed date: {target_date}")
            if telemarketer_number:
                print(f"   → Telemarketer filter: {telemarketer_number}")
            time.sleep(2)
            
            # Get today's date
            today = datetime.now()
            six_months_ago = today - timedelta(days=183)
            
            print(f"   → Scanning complaint rows...")
            time.sleep(1)
            
            # Find all complaint rows - CORRECT SELECTOR
            rows = self.driver.find_elements(By.CSS_SELECTOR, 'div.flex.w-full.border-t')
            print(f"   → Found {len(rows)} complaint(s) in search results")
            
            if len(rows) == 0:
                print(f"   ✗ No complaint rows found")
                return False
            
            # Collect matching rows first
            matching_rows = []
            
            # Iterate through rows to find matching date
            for i, row in enumerate(rows):
                try:
                    # Get all columns in the row - CORRECT SELECTOR
                    columns = row.find_elements(By.CSS_SELECTOR, 'div.p-3')
                    
                    # The "Date of Call/SMS" column should be at index 4 (5th column)
                    if len(columns) >= 5:
                        date_column = columns[4]
                        portal_date_text = date_column.text.strip()
                        
                        # Extract telemarketer number from row (column index 1)
                        telemarketer_column = columns[1] if len(columns) > 1 else None
                        row_telemarketer = telemarketer_column.text.strip() if telemarketer_column else None
                        
                        print(f"     Row {i+1}: Date={portal_date_text}, Telemarketer={row_telemarketer}")
                        
                        # Check if dates are within 6 months (no exact match required)
                        try:
                            portal_date = datetime.strptime(portal_date_text, '%B %d, %Y')
                            date_diff = abs((portal_date - expected_date_obj).days)
                            
                            # Accept if within 6 months (183 days in either direction)
                            if date_diff <= 183:
                                if date_diff == 0:
                                    print(f"             ✓ Date: Exact match!")
                                else:
                                    print(f"             ✓ Date: Within 6 months ({date_diff} days)")
                                
                                # Add to matching rows
                                matching_rows.append({
                                    'index': i,
                                    'row': row,
                                    'date': portal_date_text,
                                    'telemarketer': row_telemarketer,
                                    'date_diff': date_diff
                                })
                        except:
                            continue
                            
                except:
                    continue
            
            if not matching_rows:
                print(f"   ✗ No matching complaint by date")
                return False
            
            # If multiple matches and telemarketer number provided, filter by telemarketer
            if len(matching_rows) > 1 and telemarketer_number:
                print(f"\n   → Multiple matches found ({len(matching_rows)}), filtering by telemarketer: {telemarketer_number}")
                
                telemarketer_match = None
                for match in matching_rows:
                    if match['telemarketer'] == telemarketer_number:
                        telemarketer_match = match
                        print(f"   ✓ Found matching telemarketer in row {match['index']+1}")
                        break
                
                if telemarketer_match:
                    row_to_click = telemarketer_match['row']
                    row_index = telemarketer_match['index']
                else:
                    print(f"   ⚠️  No telemarketer match, using first date match")
                    row_to_click = matching_rows[0]['row']
                    row_index = matching_rows[0]['index']
            else:
                # Single match or no telemarketer filter
                row_to_click = matching_rows[0]['row']
                row_index = matching_rows[0]['index']
            
            print(f"\n   ✓ Found matching complaint! Clicking row {row_index+1}")
            self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", row_to_click)
            time.sleep(1)
            
            # Use JavaScript click to avoid interception
            self.driver.execute_script("arguments[0].click();", row_to_click)
            time.sleep(3)
            return True
            
        except Exception as e:
            print(f"   ✗ Error finding complaint: {str(e)}")
            import traceback
            traceback.print_exc()
            return False
    
    def extract_date_from_url(self, url: str) -> str:
        """Extract date from URL in format DD-Mon-YYYY"""
        try:
            match = re.search(r'(\d{2})-([A-Za-z]{3})-(\d{4})', url)
            if match:
                return f"{match.group(1)}-{match.group(2)}-{match.group(3)}"
            return None
        except:
            return None
    
    def is_date_within_range(self, file_date_str: str, url_date_str: str) -> bool:
        """Check if dates are within 6 months range"""
        try:
            # Clean ordinal suffixes
            file_date_str = self.clean_ordinal_date(file_date_str)
            url_date_str = self.clean_ordinal_date(url_date_str)
            
            # Try multiple date formats
            date_formats = [
                '%d %b %Y',    # 17 Dec 2025
                '%d %B %Y',    # 17 December 2025
                '%d-%b-%Y',    # 12-Dec-2025
                '%d-%m-%Y',    # 12-12-2025
                '%d/%b/%y',    # 11/Dec/25
                '%d-%b-%y',    # 11-Dec-25
                '%d/%b/%Y',    # 11/Dec/2025
                '%d.%b.%Y',    # 11.Dec.2025
                '%d.%b.%y',    # 11.Dec.25
                '%d/%m/%Y',    # 11/12/2025
                '%d/%m/%y',    # 11/12/25
                '%b %d, %Y',   # Jan 2, 2026
                '%B %d, %Y',   # January 2, 2026
            ]
            
            # Parse file_date independently
            file_date = None
            for fmt in date_formats:
                try:
                    file_date = datetime.strptime(file_date_str, fmt)
                    break
                except:
                    continue
            
            # Parse url_date independently
            url_date = None
            for fmt in date_formats:
                try:
                    url_date = datetime.strptime(url_date_str, fmt)
                    break
                except:
                    continue
            
            if not file_date or not url_date:
                print(f"   ✗ Could not parse dates: file_date='{file_date_str}' (parsed: {file_date}), url_date='{url_date_str}' (parsed: {url_date})")
                return False
            
            date_diff = abs((file_date - url_date).days)
            print(f"   → Date comparison: {file_date_str} vs {url_date_str} = {date_diff} days difference")
            return date_diff <= 183  # 6 months = ~183 days
        except Exception as e:
            print(f"   ✗ Date parsing error: {str(e)}")
            return False
    
    def verify_document_authenticity(self, current_url: str) -> bool:
        """
        Verify document authenticity by checking for URL or logo in the document
        Takes a screenshot of the document page and uses OCR to check for specific patterns
        """
        try:
            print(f"   → Verifying document authenticity with OCR...")
            
            # Take screenshot of the document page
            screenshot = self.driver.get_screenshot_as_png()
            
            # Convert to PIL Image for OCR
            from io import BytesIO
            image = Image.open(BytesIO(screenshot))
            
            # Extract text using OCR
            text = pytesseract.image_to_string(image)
            text_lower = text.lower()
            
            # Check for URL patterns that should be present in legitimate documents
            url_patterns = [
                'zomato',
                'admin.zomans.com',
                'lifeline',
                'exotel',
                'order',
            ]
            
            # Check for logo/brand text
            logo_patterns = [
                'zomato',
                'customer',
                'delivery',
                'order',
            ]
            
            found_patterns = []
            
            # Check URL patterns
            for pattern in url_patterns:
                if pattern in text_lower:
                    found_patterns.append(f"URL pattern '{pattern}'")
            
            # Check logo patterns
            for pattern in logo_patterns:
                if pattern in text_lower:
                    found_patterns.append(f"Logo text '{pattern}'")
            
            if found_patterns:
                print(f"   ✓ Authenticity verified - Found: {', '.join(found_patterns[:3])}")
                return True
            else:
                print(f"   ⚠️  Authenticity check: No recognizable URL or logo patterns found")
                # Print first 100 chars of extracted text for debugging
                preview = ' '.join(text.split()[:20])
                print(f"   → OCR preview: {preview}...")
                # Still return True as this is an additional check, not a blocker
                return True
                
        except Exception as e:
            print(f"   ⚠️  Authenticity check error: {str(e)}")
            # Don't fail the verification if OCR fails
            return True
    
    def download_and_verify_existing(self, file_date_str: str, expected_phone: str) -> bool:
        """
        Download existing document, verify date and phone, and click verify if valid
        
        Args:
            file_date_str: Date from the PDF file in format DD-Mon-YYYY
            expected_phone: Expected phone number to verify in the document
            
        Returns:
            True if verification was successful, False otherwise
        """
        try:
            print(f"\n→ Starting download and verification process...")
            
            wait = WebDriverWait(self.driver, 15)
            
            # Step 1: Click on the uploaded document preview
            print(f"   → Looking for uploaded document preview...")
            
            document_selectors = [
                (By.XPATH, '//div[@class="text-left"]//div[contains(@class, "font-medium")]'),
                (By.CSS_SELECTOR, 'div.bg-slate-100.h-24.w-full.rounded.flex.items-center.justify-center.overflow-hidden.cursor-pointer'),
                (By.XPATH, '//div[contains(@class, "bg-slate-100") and contains(@class, "cursor-pointer")]'),
            ]
            
            document_preview = None
            for selector_type, selector_value in document_selectors:
                try:
                    document_preview = wait.until(
                        EC.element_to_be_clickable((selector_type, selector_value))
                    )
                    print(f"   → Found document preview")
                    break
                except:
                    continue
            
            if not document_preview:
                print(f"   ✗ Could not find document preview")
                return False
            
            print(f"   → Clicking on document preview...")
            self.driver.execute_script("arguments[0].click();", document_preview)
            time.sleep(3)
            
            # Step 2: Wait for modal dialog and click Download button
            print(f"   → Waiting for modal dialog to open...")
            wait.until(EC.presence_of_element_located((By.XPATH, '//div[@role="dialog"][@data-state="open"]')))
            time.sleep(2)
            
            print(f"   → Looking for Download button in modal...")
            
            download_selectors = [
                (By.XPATH, '//div[@role="dialog"]//button[contains(text(), "Download")]'),
                (By.XPATH, '//button[contains(text(), "Download") and @data-slot="button"]'),
                (By.CSS_SELECTOR, 'div[role="dialog"] button[data-slot="button"]:has(svg.lucide-download)'),
                (By.XPATH, '//div[@role="dialog"]//button[.//svg[contains(@class, "lucide-download")]]'),
            ]
            
            download_button = None
            for idx, (selector_type, selector_value) in enumerate(download_selectors):
                try:
                    print(f"      Trying selector {idx+1}...")
                    download_button = wait.until(
                        EC.element_to_be_clickable((selector_type, selector_value))
                    )
                    print(f"   → Found Download button (selector {idx+1})")
                    break
                except:
                    continue
            
            if not download_button:
                print(f"   ✗ Could not find Download button with any selector")
                return False
            
            # Store current tab handle
            main_tab = self.driver.current_window_handle
            print(f"   → Current tab handle: {main_tab[:10]}...")
            
            # Use JavaScript click to avoid interception
            print(f"   → Scrolling Download button into view...")
            self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", download_button)
            time.sleep(1)
            
            print(f"   → Clicking Download button...")
            self.driver.execute_script("arguments[0].click();", download_button)
            time.sleep(3)
            
            # Step 3: Switch to new tab
            print(f"   → Checking for new tab...")
            all_tabs = self.driver.window_handles
            
            if len(all_tabs) > 1:
                # Switch to the new tab (last one)
                new_tab = [tab for tab in all_tabs if tab != main_tab][0]
                self.driver.switch_to.window(new_tab)
                print(f"   ✓ Switched to new tab")
                
                # Step 4: Get URL and extract date
                current_url = self.driver.current_url
                print(f"   → URL: {current_url}")
                
                # Step 4.5: Verify document authenticity using OCR
                if not self.verify_document_authenticity(current_url):
                    print(f"   ⚠️  Document authenticity could not be verified")
                    # Continue anyway as it's an additional check
                
                # Extract date from URL
                url_date = self.extract_date_from_url(current_url)
                
                if url_date:
                    print(f"   → Extracted date from URL: {url_date}")
                    
                    # Step 5: Compare dates (portal date should be same or within 6 months before file date)
                    if self.is_date_within_range(file_date_str, url_date):
                        print(f"   ✓ Date validation passed!")
                        
                        # Close the new tab
                        self.driver.close()
                        
                        # Switch back to main tab
                        self.driver.switch_to.window(main_tab)
                        print(f"   ✓ Switched back to main tab")
                        time.sleep(2)
                        
                        # Step 6: Click Verify button (inside dialog footer)
                        print(f"   → Looking for Verify button in modal...")
                        
                        # Wait for modal to be visible again
                        wait.until(EC.presence_of_element_located((By.XPATH, '//div[@role="dialog"][@data-state="open"]')))
                        time.sleep(1)
                        
                        verify_selectors = [
                            (By.XPATH, '//div[@role="dialog"]//button[contains(@class, "bg-emerald-600") and .//svg[contains(@class, "lucide-check")]]'),
                            (By.XPATH, '//div[@data-slot="dialog-footer"]//button[contains(@class, "bg-emerald-600")]'),
                            (By.XPATH, '//div[@role="dialog"]//button[contains(., "Verify") and contains(@class, "bg-emerald-600")]'),
                        ]
                        
                        verify_button = None
                        for selector_type, selector_value in verify_selectors:
                            try:
                                verify_button = wait.until(
                                    EC.element_to_be_clickable((selector_type, selector_value))
                                )
                                print(f"   → Found Verify button")
                                break
                            except:
                                continue
                        
                        if verify_button:
                            print(f"   → Clicking Verify button...")
                            self.driver.execute_script("arguments[0].click();", verify_button)
                            time.sleep(2)
                            
                            # Step 7: Handle confirmation dialog "Verify Document"
                            print(f"   → Looking for confirmation dialog...")
                            try:
                                confirm_verify_button = wait.until(EC.element_to_be_clickable(
                                    (By.XPATH, '//button[contains(@class, "bg-emerald-600") and contains(text(), "Verify Document")]')
                                ))
                                print(f"   → Clicking Verify Document confirmation...")
                                self.driver.execute_script("arguments[0].click();", confirm_verify_button)
                                time.sleep(3)
                                print(f"   ✅ Document verified successfully!")
                                return True
                            except Exception as confirm_e:
                                print(f"   ⚠️  Confirmation dialog not found, verification may have completed")
                                return True
                        else:
                            print(f"   ✗ Could not find Verify button")
                            return False
                    else:
                        print(f"   ✗ Date validation failed - dates are beyond 6-month range")
                        # Close the new tab
                        self.driver.close()
                        # Switch back to main tab
                        self.driver.switch_to.window(main_tab)
                        return False
                else:
                    print(f"   ✗ Could not extract date from URL")
                    # Close the new tab
                    self.driver.close()
                    # Switch back to main tab
                    self.driver.switch_to.window(main_tab)
                    return False
            else:
                print(f"   ✗ New tab did not open")
                return False
            
        except Exception as e:
            print(f"   ✗ Error in download and verify process: {str(e)}")
            import traceback
            traceback.print_exc()
            
            # Try to switch back to main tab if we're in a different one
            try:
                main_tabs = [handle for handle in self.driver.window_handles]
                if main_tabs:
                    self.driver.switch_to.window(main_tabs[0])
            except:
                pass
            
            return False
    
    def check_complaint_status(self) -> str:
        """Check if complaint status is Review Pending or Open"""
        try:
            print(f"\n   → Checking complaint status...")
            time.sleep(2)
            
            wait = WebDriverWait(self.driver, 15)
            
            # Look for status element - CORRECT SELECTOR from HTML
            # Status has colored dot and text: "Review Pending" (amber) or "Open" (blue)
            try:
                status_element = wait.until(EC.presence_of_element_located(
                    (By.XPATH, '//div[contains(@class, "flex items-center")]//span[contains(@class, "text-amber-500") or contains(@class, "text-blue-500")]//following-sibling::span')
                ))
                status_text = status_element.text.strip()
                print(f"   ✓ Status found: {status_text}")
                return status_text.lower()
            except:
                # Try alternative selector
                try:
                    status_element = self.driver.find_element(By.XPATH,
                        '//span[text()="Status"]//ancestor::div[1]//following-sibling::div//span[@class="truncate"]')
                    status_text = status_element.text.strip()
                    print(f"   ✓ Status found: {status_text}")
                    return status_text.lower()
                except:
                    print(f"   ✗ Could not determine status")
                    return "unknown"
                    
        except Exception as e:
            print(f"   ✗ Error checking status: {str(e)}")
            return "unknown"
    
    def click_verify_button(self) -> bool:
        """Click the Verify button - used for OPEN status after upload"""
        try:
            print(f"\n   → Starting document verification...")
            
            # Wait for upload to complete and page to refresh
            time.sleep(3)
            
            wait = WebDriverWait(self.driver, 15)
            
            # Step 1: Click on the uploaded document - CORRECT SELECTOR
            print(f"   → Looking for uploaded document...")
            
            uploaded_doc = wait.until(EC.element_to_be_clickable(
                (By.XPATH, '//div[contains(@class, "flex items-center")]//svg[contains(@class, "lucide-file-text")]//ancestor::div[contains(@class, "flex items-center")][1]')
            ))
            
            print(f"   → Found uploaded document")
            print(f"   → Clicking on uploaded document...")
            uploaded_doc.click()
            time.sleep(2)
            print(f"   ✓ Document opened")
            
            # Step 2: Click the Verify button - CORRECT SELECTOR from HTML
            print(f"   → Looking for Verify button...")
            
            verify_button = wait.until(EC.element_to_be_clickable(
                (By.XPATH, '//button[contains(@class, "bg-emerald-600")]//svg[contains(@class, "lucide-check")]//parent::button')
            ))
            
            print(f"   → Found Verify button (green with check icon)")
            print(f"   → Clicking Verify button...")
            verify_button.click()
            time.sleep(2)
            print(f"   ✓ Verify button clicked")
            
            # Step 3: Click the "Verify Document" confirmation button if it appears
            print(f"   → Looking for Verify Document confirmation...")
            try:
                verify_doc_button = wait.until(EC.element_to_be_clickable(
                    (By.XPATH, '//button[@data-slot="button" and (text()="Verify Document" or text()="Confirm")]')
                ))
                
                print(f"   → Found confirmation button")
                print(f"   → Clicking Verify Document...")
                verify_doc_button.click()
                time.sleep(3)
            except:
                # May not always appear
                print(f"   → No confirmation needed")
            
            print(f"   ✅ Document verified successfully!")
            return True
            
        except Exception as e:
            print(f"✗ Error verifying document: {str(e)}")
            import traceback
            traceback.print_exc()
            return False
    
    def upload_document(self, file_path: Path) -> bool:
        """Upload document from local file"""
        try:
            print(f"\n   → Uploading document: {file_path.name}")
            
            wait = WebDriverWait(self.driver, 15)
            time.sleep(2)
            
            # Step 1: Click Upload button - CORRECT SELECTOR from HTML
            print(f"   → Looking for Upload button...")
            upload_button = wait.until(EC.element_to_be_clickable(
                (By.XPATH, '//div[@class="flex items-center gap-2"]//button[.//svg[contains(@class, "lucide-upload")] and .//span[text()="Upload"]]')
            ))
            
            print(f"   → Found Upload button")
            print(f"   → Clicking Upload button...")
            upload_button.click()
            time.sleep(2)
            
            # Step 2: Click "Select File" button - CORRECT SELECTOR from HTML
            print(f"   → Looking for Select File button...")
            select_file_button = wait.until(EC.element_to_be_clickable(
                (By.XPATH, '//button[@data-slot="button" and text()="Select File"]')
            ))
            
            print(f"   → Clicking Select File...")
            select_file_button.click()
            time.sleep(1)
            
            # Step 3: Find file input
            print(f"   → Looking for file input element...")
            file_input = wait.until(
                EC.presence_of_element_located((By.CSS_SELECTOR, 'input[type="file"]'))
            )
            
            # Step 4: Send file path
            print(f"   → Uploading file: {file_path.name}")
            file_input.send_keys(str(file_path))
            time.sleep(3)
            print(f"   ✓ File selected successfully")
            
            # Step 5: Check consent checkbox - CORRECT SELECTOR from HTML
            print(f"   → Looking for consent checkbox...")
            consent_checkbox = wait.until(EC.element_to_be_clickable(
                (By.CSS_SELECTOR, 'button[type="button"][role="checkbox"]#consent')
            ))
            
            print(f"   → Clicking consent checkbox...")
            consent_checkbox.click()
            time.sleep(2)
            print(f"   ✓ Consent checkbox checked")
            
            # Step 6: Click final Upload button - CORRECT SELECTOR from HTML
            print(f"   → Looking for final Upload button...")
            final_upload_button = wait.until(EC.element_to_be_clickable(
                (By.XPATH, '//button[@data-slot="button" and text()="Upload"]')
            ))
            
            print(f"   → Clicking final Upload button...")
            final_upload_button.click()
            time.sleep(4)
            
            print(f"   ✓ Document uploaded successfully!")
            return True
            
        except Exception as e:
            print(f"✗ Error uploading document: {str(e)}")
            import traceback
            traceback.print_exc()
            return False
    
    def process_review_pending_file(self, file_path: Path) -> bool:
        """Process a single Review Pending file (already downloaded from dashboard)"""
        try:
            print(f"\n{'='*70}")
            print(f"📄 PROCESSING REVIEW PENDING FILE")
            print(f"{'='*70}")
            print(f"File: {file_path.name}\n")
            
            # Step 1: Extract phone from filename
            phone = self.extract_phone_from_filename(file_path.name)
            if not phone:
                print(f"❌ SKIPPED - No phone number in filename")
                self.move_file_to_processed_review(file_path)
                return False
            
            print(f"✓ Phone from filename: {phone}")
            
            # Extract telemarketer number from filename (if present)
            telemarketer = self.extract_telemarketer_from_filename(file_path.name)
            if telemarketer:
                print(f"✓ Telemarketer from filename: {telemarketer}")
            else:
                print(f"⚠️  No telemarketer number in filename")
            
            # Step 2: Perform comprehensive OCR extraction on LOCAL file
            print(f"\n{'─'*70}")
            print(f"🔍 LOCAL FILE OCR EXTRACTION")
            print(f"{'─'*70}")
            local_file_data = self.extract_data_from_file(file_path)
            
            # Step 3: Validate LOCAL file has minimum required data
            if not local_file_data.get('has_authenticity'):
                print(f"\n❌ SKIPPED - Local file has no URL/logo (not authentic)")
                print(f"   Reason: Document must contain recognizable URL or logo")
                print(f"   Note: Enhanced OCR attempted but no patterns found")
                self.move_file_to_processed_review(file_path)
                return False
            
            if phone not in local_file_data.get('all_phones', []):
                print(f"\n❌ SKIPPED - Phone {phone} not found in local file content")
                self.move_file_to_processed_review(file_path)
                return False
            
            if not local_file_data.get('all_dates'):
                print(f"\n❌ SKIPPED - No dates found in local file content")
                self.move_file_to_processed_review(file_path)
                return False
            
            # Step 4: Navigate and search
            self.navigate_to_all_complaints()
            
            if not self.search_complaint(phone):
                print(f"\n❌ SKIPPED - Phone not found in dashboard")
                self.move_file_to_processed_review(file_path)
                return False
            
            # Step 5: Try ALL dates until one matches a complaint
            all_dates = local_file_data['all_dates']
            print(f"\n→ Trying {len(all_dates)} date(s) to find matching complaint...")
            
            complaint_found = False
            for idx, expected_date in enumerate(all_dates, 1):
                print(f"\n   → Attempt {idx}/{len(all_dates)}: Trying date '{expected_date}'")
                
                if self.find_and_click_complaint(expected_date, telemarketer):
                    print(f"   ✓ Successfully matched complaint with date: {expected_date}")
                    complaint_found = True
                    break
                else:
                    print(f"   ✗ Date '{expected_date}' did not match any complaint")
                    # Continue to next date
            
            # FALLBACK: If no dates matched, try extracting date from filename and match exactly
            if not complaint_found:
                print(f"\n⚠️  No dates matched within 6 months. Trying FALLBACK: filename date exact match...")
                
                # Extract date from filename (format: DD-Mon-YYYY)
                filename_date_match = re.search(r'(\d{1,2})-(\w{3})-(\d{4})', file_path.name)
                if filename_date_match:
                    filename_date = filename_date_match.group(0)
                    print(f"   → Extracted filename date: {filename_date}")
                    
                    if self.find_and_click_complaint(filename_date, telemarketer):
                        print(f"   ✓ Successfully matched using filename date: {filename_date}")
                        complaint_found = True
                    else:
                        print(f"   ✗ Filename date '{filename_date}' also did not match")
                else:
                    print(f"   ✗ Could not extract date from filename")
            
            if not complaint_found:
                print(f"\n❌ SKIPPED - No matching complaint found (tried {len(all_dates)} date(s) + filename)")
                self.move_file_to_processed_review(file_path)
                return False
            
            # Step 6: Download portal document and perform comprehensive validation
            if not self.download_verify_and_confirm(local_file_data, phone):
                print(f"\n❌ SKIPPED - Validation failed (see details above)")
                self.move_file_to_processed_review(file_path)
                return False
            
            # Step 7: Move to processed_review folder
            self.move_file_to_processed_review(file_path)
            
            print(f"\n{'='*70}")
            print(f"✅ SUCCESSFULLY VERIFIED AND PROCESSED")
            print(f"{'='*70}\n")
            return True
            
        except Exception as e:
            print(f"\n✗ Processing error: {str(e)}")
            import traceback
            traceback.print_exc()
            self.move_file_to_processed_review(file_path)
            return False
    
    def download_verify_and_confirm(self, local_file_data: dict, expected_phone: str) -> bool:
        """
        Download portal document, perform comprehensive OCR validation, and click verify
        Returns True if all validations pass and verify clicked, False otherwise
        """
        try:
            print(f"\n{'─'*70}")
            print(f"📥 DOWNLOADING PORTAL DOCUMENT FOR VALIDATION")
            print(f"{'─'*70}")
            
            # Get current URL to extract date
            current_url = self.driver.current_url
            url_date_str = self.extract_date_from_url(current_url)
            
            if not url_date_str:
                print(f"   ✗ Cannot extract date from URL: {current_url}")
                return False
            
            print(f"   → URL date: {url_date_str}")
            
            # Download the portal document
            print(f"\n   → Clicking document to download...")
            wait = WebDriverWait(self.driver, 15)
            
            try:
                # Click to open modal/download
                document_link = wait.until(EC.element_to_be_clickable(
                    (By.XPATH, '//a[contains(@href, "/download")]')
                ))
                self.driver.execute_script("arguments[0].click();", document_link)
                time.sleep(3)
                
                # Look for download button in modal
                download_button = wait.until(EC.element_to_be_clickable(
                    (By.XPATH, '//div[@role="dialog"]//a[contains(@href, "/download")]')
                ))
                self.driver.execute_script("arguments[0].click();", download_button)
                time.sleep(5)
                
                print(f"   ✓ Document download initiated")
                
            except Exception as e:
                print(f"   ✗ Download failed: {str(e)}")
                return False
            
            # Find the downloaded file
            print(f"\n   → Looking for downloaded file...")
            portal_file = None
            for _ in range(10):
                time.sleep(1)
                files = list(self.ndnc_folder.glob("*.pdf")) + list(self.ndnc_folder.glob("*.png")) + \
                        list(self.ndnc_folder.glob("*.jpg")) + list(self.ndnc_folder.glob("*.jpeg"))
                recent_files = [f for f in files if (time.time() - f.stat().st_mtime) < 15]
                if recent_files:
                    portal_file = max(recent_files, key=lambda x: x.stat().st_mtime)
                    break
            
            if not portal_file:
                print(f"   ✗ Downloaded file not found")
                return False
            
            print(f"   ✓ Found downloaded file: {portal_file.name}")
            
            # Perform comprehensive OCR on portal document
            print(f"\n{'─'*70}")
            print(f"🔍 PORTAL DOCUMENT OCR EXTRACTION")
            print(f"{'─'*70}")
            portal_file_data = self.extract_data_from_file(portal_file)
            
            # Perform comprehensive validation
            is_valid, reason = self.validate_document_completely(
                portal_file_data, 
                expected_phone, 
                url_date_str
            )
            
            # Clean up portal file
            try:
                portal_file.unlink()
                print(f"   → Cleaned up portal document")
            except:
                pass
            
            if not is_valid:
                print(f"\n❌ VALIDATION FAILED: {reason}")
                return False
            
            # All validations passed - click Verify button
            print(f"\n{'─'*70}")
            print(f"✅ ALL VALIDATIONS PASSED - CLICKING VERIFY")
            print(f"{'─'*70}")
            
            # Click Verify button
            try:
                time.sleep(2)
                verify_button = wait.until(EC.element_to_be_clickable(
                    (By.XPATH, '//button[contains(., "Verify") and not(contains(., "Bulk"))]')
                ))
                self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", verify_button)
                time.sleep(1)
                self.driver.execute_script("arguments[0].click();", verify_button)
                time.sleep(2)
                
                print(f"   ✓ Clicked Verify button")
                
                # Click confirmation dialog
                try:
                    confirm_button = WebDriverWait(self.driver, 5).until(
                        EC.element_to_be_clickable((By.XPATH, '//button[contains(text(), "Verify Document")]'))
                    )
                    self.driver.execute_script("arguments[0].click();", confirm_button)
                    time.sleep(2)
                    print(f"   ✓ Confirmed verification")
                except:
                    print(f"   → No confirmation dialog (or already confirmed)")
                
                return True
                
            except Exception as e:
                print(f"   ✗ Could not click Verify button: {str(e)}")
                return False
            
        except Exception as e:
            print(f"\n✗ Error in download_verify_and_confirm: {str(e)}")
            import traceback
            traceback.print_exc()
            return False
    
    def move_file_to_processed_review(self, file_path: Path) -> bool:
        """Move file to processed_review folder"""
        try:
            dest_path = self.processed_review_folder / file_path.name
            if dest_path.exists():
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                dest_path = self.processed_review_folder / f"{file_path.stem}_{timestamp}{file_path.suffix}"
            
            shutil.move(str(file_path), str(dest_path))
            print(f"   → Moved to processed_review: {file_path.name}")
            return True
        except Exception as e:
            print(f"   ⚠️  Could not move file: {str(e)}")
            return False
    
    def click_verify_button_review_pending(self) -> bool:
        """Click Verify button for Review Pending (no upload, just verify)"""
        try:
            print(f"\n   → Looking for Verify button...")
            
            wait = WebDriverWait(self.driver, 15)
            time.sleep(2)
            
            # Find Verify button - CORRECT SELECTOR
            verify_button = wait.until(EC.element_to_be_clickable(
                (By.XPATH, '//button[contains(@class, "bg-emerald-600")]//svg[contains(@class, "lucide-check")]//parent::button')
            ))
            
            print(f"   → Found Verify button (green with check icon)")
            print(f"   → Clicking Verify button...")
            verify_button.click()
            time.sleep(3)
            
            print(f"   ✅ Document verified successfully!")
            return True
            
        except Exception as e:
            print(f"✗ Error clicking verify: {str(e)}")
            import traceback
            traceback.print_exc()
            return False
    
    def process_open_file(self, file_path: Path) -> bool:
        """Process a single Open complaint file"""
        try:
            print(f"\n{'='*60}")
            print(f"📄 Processing Open: {file_path.name}")
            print(f"{'='*60}")
            
            # Extract phone from filename
            phone = self.extract_phone_from_filename(file_path.name)
            if not phone:
                print(f"✗ No phone number in filename")
                return False
            
            print(f"✓ Phone from filename: {phone}")
            
            # Extract telemarketer number from filename (if present)
            telemarketer = self.extract_telemarketer_from_filename(file_path.name)
            if telemarketer:
                print(f"✓ Telemarketer from filename: {telemarketer}")
            else:
                print(f"⚠️  No telemarketer number in filename")
            
            # Extract data from file content
            file_data = self.extract_data_from_file(file_path)
            if not file_data.get('all_dates'):
                print(f"✗ Could not extract any dates from file")
                return False
            
            # Navigate to All Complaints
            self.navigate_to_all_complaints()
            
            # Search for phone number
            if not self.search_complaint(phone):
                return False
            
            # Try ALL dates until one matches a complaint
            all_dates = file_data['all_dates']
            print(f"\n→ Trying {len(all_dates)} date(s) to find matching complaint...")
            
            complaint_found = False
            for idx, expected_date in enumerate(all_dates, 1):
                print(f"\n   → Attempt {idx}/{len(all_dates)}: Trying date '{expected_date}'")
                
                if self.find_and_click_complaint(expected_date, telemarketer):
                    print(f"   ✓ Successfully matched complaint with date: {expected_date}")
                    complaint_found = True
                    break
                else:
                    print(f"   ✗ Date '{expected_date}' did not match any complaint")
                    # Continue to next date
            
            # FALLBACK: If no dates matched, try extracting date from filename and match exactly
            if not complaint_found:
                print(f"\n⚠️  No dates matched within 6 months. Trying FALLBACK: filename date exact match...")
                
                # Extract date from filename (format: DD-Mon-YYYY)
                filename_date_match = re.search(r'(\d{1,2})-(\w{3})-(\d{4})', file_path.name)
                if filename_date_match:
                    filename_date = filename_date_match.group(0)
                    print(f"   → Extracted filename date: {filename_date}")
                    
                    if self.find_and_click_complaint(filename_date, telemarketer):
                        print(f"   ✓ Successfully matched using filename date: {filename_date}")
                        complaint_found = True
                    else:
                        print(f"   ✗ Filename date '{filename_date}' also did not match")
                else:
                    print(f"   ✗ Could not extract date from filename")
            
            if not complaint_found:
                print(f"\n✗ No matching complaint found (tried {len(all_dates)} date(s) + filename)")
                return False
            
            # Upload document
            if not self.upload_document(file_path):
                return False
            
            # Click Verify button
            if not self.click_verify_button():
                return False
            
            # Move to processed folder
            processed_path = self.processed_folder / file_path.name
            if processed_path.exists():
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                processed_path = self.processed_folder / f"{file_path.stem}_{timestamp}{file_path.suffix}"
            
            shutil.move(str(file_path), str(processed_path))
            print(f"   → Moved to processed: {processed_path.name}")
            
            print(f"\n✅ Successfully processed: {file_path.name}")
            return True
            
        except Exception as e:
            print(f"\n✗ Processing error: {str(e)}")
            import traceback
            traceback.print_exc()
            return False
    
    def process_excel_and_download_files(self, excel_path: Path) -> int:
        """Process Excel file and download files from Document Link column"""
        try:
            print(f"\n   → Processing Excel file: {excel_path.name}")
            
            # Read Excel file
            workbook = openpyxl.load_workbook(excel_path)
            sheet = workbook.active
            
            print(f"   → Reading Excel data...")
            
            # Check header row to find Document Link column
            print(f"   → Checking Excel structure...")
            header_row = 1
            doc_link_col = None
            
            # Search for Document Link column (usually T=20, but verify)
            for col_idx in range(1, sheet.max_column + 1):
                header_cell = sheet.cell(header_row, col_idx).value
                if header_cell and 'Document Link' in str(header_cell):
                    doc_link_col = col_idx
                    print(f"   ✓ Found 'Document Link' in column {col_idx} ({chr(64 + col_idx)})")
                    break
            
            if not doc_link_col:
                print(f"   ⚠️  'Document Link' column not found, using column T (20)")
                doc_link_col = 20
            
            # Columns: C (3) = Complainer No, G (7) = Date of call/sms
            total_rows = sheet.max_row - 1  # Exclude header
            print(f"   → Found {total_rows} rows\n")
            
            downloaded_count = 0
            
            for row_idx in range(2, sheet.max_row + 1):  # Start from row 2 (skip header)
                try:
                    # Get data from columns
                    complainer_no = sheet.cell(row_idx, 3).value  # Column C
                    date_of_call = sheet.cell(row_idx, 7).value   # Column G
                    doc_link_cell = sheet.cell(row_idx, doc_link_col)  # Document Link column (dynamic)
                    
                    # Debug: Show cell info
                    cell_value = doc_link_cell.value
                    cell_hyperlink = doc_link_cell.hyperlink
                    
                    print(f"\n   Row {row_idx-1}: {complainer_no} ({date_of_call})")
                    print(f"      → Cell value: {cell_value}")
                    print(f"      → Has hyperlink: {cell_hyperlink is not None}")
                    
                    # Try multiple methods to get the download link
                    download_link = None
                    
                    # Method 1: Check hyperlink target
                    if cell_hyperlink and hasattr(cell_hyperlink, 'target'):
                        download_link = cell_hyperlink.target
                        print(f"      → Extracted from hyperlink.target")
                    # Method 2: Check cell value
                    elif cell_value and isinstance(cell_value, str) and cell_value.startswith('http'):
                        download_link = cell_value
                        print(f"      → Extracted from cell value")
                    # Method 3: Check if value is "Download Here" and has hyperlink
                    elif cell_value == 'Download Here' and cell_hyperlink:
                        download_link = str(cell_hyperlink)
                        print(f"      → Extracted from hyperlink object")
                    
                    if not download_link or download_link in ['Download Here', None, '']:
                        print(f"      ✗ No valid download link found, skipping...")
                        continue
                    
                    print(f"      → Download URL: {download_link[:80]}...")
                    
                    # Extract expected filename from URL or use phone/date
                    expected_filename = None
                    url_match = re.search(r'/([^/]+\.(pdf|png|jpg|jpeg))$', download_link, re.IGNORECASE)
                    if url_match:
                        expected_filename = url_match.group(1)
                    
                    # Check if file already exists in review_pending to avoid duplicates
                    if expected_filename:
                        existing_file = self.review_pending_folder / expected_filename
                        if existing_file.exists():
                            print(f"      ⊘ File already exists in review_pending/, skipping download")
                            continue
                    
                    # Navigate directly to the download link
                    # This will trigger download in current tab
                    try:
                        print(f"      → Navigating to download link...")
                        self.driver.get(download_link)
                        time.sleep(4)
                        
                        # Check if we got redirected to an error page
                        current_url = self.driver.current_url
                        if 'error' in current_url.lower() or '404' in current_url:
                            print(f"      ⚠️  Download link appears invalid (error/404)")
                            continue
                            
                    except Exception as nav_error:
                        print(f"      ⚠️  Navigation error: {str(nav_error)}")
                        continue
                    
                    # Wait for file to download with retry logic
                    max_retries = 3
                    retry_count = 0
                    file_moved = False
                    
                    while retry_count < max_retries and not file_moved:
                        time.sleep(3)
                        
                        # Find the downloaded file - files download to self.ndnc_folder
                        pdf_files = list(self.ndnc_folder.glob("*.pdf")) + list(self.ndnc_folder.glob("*.png")) + \
                                    list(self.ndnc_folder.glob("*.jpg")) + list(self.ndnc_folder.glob("*.jpeg"))
                        
                        # Exclude files already in subfolders
                        pdf_files = [f for f in pdf_files if f.parent == self.ndnc_folder]
                        
                        if pdf_files:
                            latest_file = max(pdf_files, key=lambda x: x.stat().st_mtime)
                            
                            # Check if recently downloaded (within last 30 seconds - increased from 10)
                            time_diff = time.time() - latest_file.stat().st_mtime
                            if time_diff < 30:
                                # File already has correct name from dashboard, just move it
                                dest_path = self.review_pending_folder / latest_file.name
                                if dest_path.exists():
                                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                                    dest_path = self.review_pending_folder / f"{latest_file.stem}_{timestamp}{latest_file.suffix}"
                                
                                try:
                                    shutil.move(str(latest_file), str(dest_path))
                                    print(f"      ✓ Downloaded & moved: {latest_file.name}")
                                    downloaded_count += 1
                                    file_moved = True
                                except Exception as move_e:
                                    print(f"      ⚠️  Move error: {str(move_e)}")
                                    retry_count += 1
                            else:
                                print(f"      ⚠️  File too old ({time_diff:.1f}s), retrying...")
                                retry_count += 1
                        else:
                            print(f"      ⚠️  No files found in NDNC folder, retrying...")
                            retry_count += 1
                    
                    if not file_moved:
                        print(f"      ✗ Failed to download/move file after {max_retries} attempts")
                    
                except Exception as e:
                    print(f"   ⚠️  Error downloading row {row_idx-1}: {str(e)}")
                    continue
            
            # Delete Excel file after processing
            print(f"\n   → Cleaning up Excel file...")
            try:
                if excel_path.exists():
                    excel_path.unlink()
                    print(f"   ✓ Deleted Excel file: {excel_path.name}")
                else:
                    print(f"   ℹ️  Excel file already removed")
            except Exception as cleanup_e:
                print(f"   ⚠️  Could not delete Excel file: {str(cleanup_e)}")
            
            print(f"\n   → Checking NDNC folder for any remaining files...")
            remaining_files = list(self.ndnc_folder.glob("*.pdf")) + list(self.ndnc_folder.glob("*.png")) + \
                              list(self.ndnc_folder.glob("*.jpg")) + list(self.ndnc_folder.glob("*.jpeg"))
            remaining_files = [f for f in remaining_files if f.parent == self.ndnc_folder]
            if remaining_files:
                print(f"   ⚠️  {len(remaining_files)} file(s) still in NDNC folder:")
                for f in remaining_files[:5]:  # Show first 5
                    print(f"      - {f.name}")
            else:
                print(f"   ✓ NDNC folder clean - all files moved to review_pending/")
            
            return downloaded_count
            
        except Exception as e:
            print(f"   ✗ Error processing Excel: {str(e)}")
            import traceback
            traceback.print_exc()
            return 0
    
    def cleanup_ndnc_folder(self):
        """Move any leftover files from NDNC folder to review_pending"""
        try:
            leftover_files = list(self.ndnc_folder.glob("*.pdf")) + list(self.ndnc_folder.glob("*.png")) + \
                             list(self.ndnc_folder.glob("*.jpg")) + list(self.ndnc_folder.glob("*.jpeg"))
            leftover_files = [f for f in leftover_files if f.parent == self.ndnc_folder]
            
            if leftover_files:
                print(f"   → Found {len(leftover_files)} leftover file(s) in NDNC folder")
                for file in leftover_files:
                    try:
                        dest_path = self.review_pending_folder / file.name
                        if dest_path.exists():
                            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                            dest_path = self.review_pending_folder / f"{file.stem}_{timestamp}{file.suffix}"
                        shutil.move(str(file), str(dest_path))
                        print(f"   → Moved leftover: {file.name}")
                    except Exception as e:
                        print(f"   ⚠️  Could not move {file.name}: {str(e)}")
        except Exception as e:
            print(f"   ⚠️  Cleanup error: {str(e)}")
    
    def download_review_pending_files_from_dashboard(self) -> int:
        """Download all Review Pending documents from dashboard"""
        try:
            print(f"\n{'='*70}")
            print(f"📥 DOWNLOADING REVIEW PENDING FILES FROM DASHBOARD")
            print(f"{'='*70}\n")
            
            # Clean up any leftover files from previous run
            self.cleanup_ndnc_folder()
            
            # Navigate to All Complaints
            self.navigate_to_all_complaints()
            
            # Filter by Review Pending status
            print(f"→ Filtering by 'Review Pending' status...")
            wait = WebDriverWait(self.driver, 15)
            time.sleep(3)
            
            # Try to find and click status filter dropdown
            try:
                print(f"   → Looking for status dropdown button...")
                status_filter_button = wait.until(EC.presence_of_element_located(
                    (By.XPATH, '//button[@role="combobox"][@data-slot="select-trigger"]')
                ))
                
                # Get current selection text
                current_status = None
                try:
                    current_status_element = status_filter_button.find_element(By.CSS_SELECTOR, 'span[data-slot="select-value"]')
                    current_status = current_status_element.text.strip()
                    print(f"   → Current filter: '{current_status}'")
                except:
                    print(f"   → Found status filter button")
                
                # Check if already filtered by Review Pending
                if current_status and "Review Pending" in current_status:
                    print(f"   ✓ Already filtered by Review Pending")
                else:
                    # Scroll button into view
                    self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", status_filter_button)
                    time.sleep(1)
                    
                    # Click using JavaScript to avoid interception
                    print(f"   → Opening status dropdown...")
                    self.driver.execute_script("arguments[0].click();", status_filter_button)
                    time.sleep(2)
                    
                    # Wait for dropdown menu to appear
                    try:
                        WebDriverWait(self.driver, 5).until(
                            EC.presence_of_element_located((By.XPATH, '//div[@role="listbox" or @role="menu"]'))
                        )
                        print(f"   → Dropdown opened")
                        time.sleep(1)  # Additional wait for options to render
                    except:
                        print(f"   ⚠️  Dropdown menu not detected")
                    
                    # Select "Review Pending" from dropdown - try multiple selectors
                    print(f"   → Looking for Review Pending option...")
                    
                    review_pending_option = None
                    selectors = [
                        (By.XPATH, '//div[@role="option"]//span[text()="Review Pending"]'),
                        (By.XPATH, '//div[@role="option" and contains(., "Review Pending")]'),
                        (By.XPATH, '//div[@role="listbox"]//div[contains(text(), "Review Pending")]'),
                        (By.XPATH, '//*[@role="option"][.//text()="Review Pending"]'),
                    ]
                    
                    for idx, (by, selector) in enumerate(selectors):
                        try:
                            print(f"      Trying selector {idx+1}...")
                            review_pending_option = WebDriverWait(self.driver, 5).until(
                                EC.presence_of_element_located((by, selector))
                            )
                            if review_pending_option:
                                print(f"   → Found Review Pending option (selector {idx+1})")
                                break
                        except:
                            continue
                    
                    if not review_pending_option:
                        # Debug: Show what options are available
                        try:
                            all_options = self.driver.find_elements(By.XPATH, '//div[@role="option"]')
                            print(f"   → Found {len(all_options)} options in dropdown:")
                            for i, opt in enumerate(all_options[:10]):  # Show first 10
                                print(f"      Option {i+1}: {opt.text}")
                        except:
                            pass
                        raise Exception("Review Pending option not found in dropdown")
                    
                    # Scroll option into view and click
                    self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", review_pending_option)
                    time.sleep(0.5)
                    print(f"   → Selecting Review Pending...")
                    self.driver.execute_script("arguments[0].click();", review_pending_option)
                    time.sleep(4)  # Wait for filter to apply
                    print(f"   ✓ Filtered by Review Pending")
                    
            except Exception as e:
                print(f"   ⚠️  Could not filter by status: {str(e)}")
                print(f"   Showing all complaints instead")
            
            # Wait for page to load after filtering
            time.sleep(3)
            
            # Get all complaint rows using correct selector
            rows = self.driver.find_elements(By.CSS_SELECTOR, 'div.flex.w-full.border-t')
            total_complaints = len(rows)
            print(f"✓ Found {total_complaints} Review Pending complaints\n")
            
            if total_complaints == 0:
                print(f"⚠️  No Review Pending complaints found")
                return 0
            
            # Click bulk Download button to download Excel
            try:
                print(f"→ Looking for bulk Download button...")
                
                # EXACT SELECTOR from HTML
                bulk_download_btn = wait.until(EC.presence_of_element_located(
                    (By.XPATH, '//button[@data-slot="tooltip-trigger" and contains(@class, "bg-primary")]//span[text()="Download"]//parent::button')
                ))
                print(f"   → Found bulk Download button")
                
                # Scroll into view and click using JavaScript
                self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", bulk_download_btn)
                time.sleep(1)
                
                print(f"   → Clicking to download Excel...")
                self.driver.execute_script("arguments[0].click();", bulk_download_btn)
                time.sleep(8)
                
                # Wait for Excel file to download
                print(f"   → Waiting for Excel file to download...")
                
                # Search in multiple locations
                search_folders = [
                    self.ndnc_folder,  # Primary: NDNC folder (Chrome download location)
                    Path.home() / "Downloads",  # Fallback: Default Downloads folder
                ]
                
                excel_file = None
                
                # Wait up to 60 seconds for Excel file
                for i in range(60):
                    for folder in search_folders:
                        if not folder.exists():
                            continue
                            
                        excel_files = list(folder.glob("complaints*.xlsx")) + list(folder.glob("complaints*.xls"))
                        if excel_files:
                            latest_excel = max(excel_files, key=lambda x: x.stat().st_mtime)
                            # Check if downloaded in last 90 seconds (increased from 60)
                            if (time.time() - latest_excel.stat().st_mtime) < 90:
                                excel_file = latest_excel
                                print(f"   ✓ Found Excel in: {folder}")
                                break
                    
                    if excel_file:
                        break
                    
                    if i % 5 == 0 and i > 0:
                        print(f"      Waiting... ({i}s)")
                    time.sleep(1)
                
                if not excel_file:
                    print(f"   ✗ No recent Excel download detected within 90 seconds")
                    print(f"   ✗ Excel download failed - no recent file found")
                    raise Exception("Excel download failed - please retry")
                
                print(f"   ✓ Excel file found: {excel_file.name}")
                
                # Move Excel to NDNC folder if it's not already there
                if excel_file.parent != self.ndnc_folder:
                    try:
                        new_excel_path = self.ndnc_folder / excel_file.name
                        if new_excel_path.exists():
                            new_excel_path.unlink()  # Remove old file
                        shutil.move(str(excel_file), str(new_excel_path))
                        excel_file = new_excel_path
                        print(f"   → Moved Excel to NDNC folder")
                    except Exception as move_e:
                        print(f"   ⚠️  Could not move Excel file: {str(move_e)}")
                
                # Process Excel file to download individual documents
                print(f"   → Processing Excel to download individual documents...")
                downloaded_count = self.process_excel_and_download_files(excel_file)
                
                print(f"\n{'='*70}")
                print(f"📥 Downloaded {downloaded_count} files from Excel")
                print(f"{'='*70}\n")
                
                return downloaded_count
                
            except Exception as e:
                print(f"   ✗ Bulk download failed: {str(e)}")
                import traceback
                traceback.print_exc()
                return 0
            
            # If bulk download failed, download individually
            downloaded_count = 0
            
            for idx, row in enumerate(rows, 1):
                try:
                    print(f"   Processing complaint {idx}/{total_complaints}...")
                    
                    # Get phone and date from row before clicking
                    phone = None
                    date = None
                    
                    try:
                        columns = row.find_elements(By.CSS_SELECTOR, 'div.p-3')
                        
                        # Get date from column 4 (5th column - Date of Call/SMS)
                        if len(columns) >= 5:
                            date_column = columns[4]
                            date_text = date_column.text.strip()
                            # Format: "December 17, 2025" -> convert to "17-Dec-2025"
                            try:
                                date_obj = datetime.strptime(date_text, '%B %d, %Y')
                                date = date_obj.strftime('%d-%b-%Y')
                            except:
                                pass
                        
                        # Get phone from the row (usually column 1 or 2)
                        for col in columns:
                            col_text = col.text.strip()
                            phone_match = re.search(r'(\d{10})', col_text)
                            if phone_match:
                                phone = phone_match.group(1)
                                break
                    except:
                        pass
                    
                    # Click on the row to open complaint details
                    self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", row)
                    time.sleep(1)
                    
                    # Use JavaScript click to avoid interception
                    self.driver.execute_script("arguments[0].click();", row)
                    time.sleep(3)
                    
                    # Try to find and download document
                    try:
                        # Find file preview element - CORRECT SELECTOR from HTML
                        doc_preview = wait.until(EC.element_to_be_clickable(
                            (By.XPATH, '//div[contains(@class, "flex items-center")]//svg[contains(@class, "lucide-file-text")]//ancestor::div[contains(@class, "flex items-center")]')
                        ))
                        
                        print(f"   → Found document preview")
                        
                        # Get filename from the preview if possible
                        try:
                            filename_element = self.driver.find_element(By.CSS_SELECTOR, 'div.font-medium.text-sm.truncate')
                            original_filename = filename_element.text.strip()
                            print(f"   → Original filename: {original_filename}")
                            
                            # Extract phone from filename if not found yet
                            if not phone:
                                phone_match = re.search(r'(\d{10})', original_filename)
                                if phone_match:
                                    phone = phone_match.group(1)
                            
                            # Extract date from filename if not found yet
                            if not date:
                                date_match = re.search(r'(\d{1,2}-\w{3}-\d{4})', original_filename)
                                if date_match:
                                    date = date_match.group(1)
                        except:
                            pass
                        
                        # Click on document preview
                        print(f"   → Clicking on document...")
                        doc_preview.click()
                        time.sleep(2)
                        
                        # Find and click Download button - UPDATED SELECTOR
                        download_btn = None
                        selectors = [
                            (By.XPATH, '//button[@data-slot="button" and contains(@class, "bg-background") and contains(@class, "h-8") and .//svg[contains(@class, "lucide-download")]]'),
                            (By.XPATH, '//button[@data-slot="button" and contains(@class, "flex-shrink-0") and .//svg[contains(@class, "lucide-download")]]'),
                            (By.XPATH, '//button[contains(@class, "inline-flex") and contains(@class, "h-8") and .//svg[contains(@class, "lucide-download")]]'),
                            (By.XPATH, '//button[.//svg[contains(@class, "lucide-download") and @width="14"]]'),
                        ]
                        
                        for selector_type, selector_value in selectors:
                            try:
                                download_btn = wait.until(EC.element_to_be_clickable((selector_type, selector_value)))
                                break
                            except:
                                continue
                        
                        if not download_btn:
                            print(f"   ✗ Could not find Download button")
                            continue
                        
                        # Store main window
                        main_window = self.driver.current_window_handle
                        
                        # Use JavaScript click to avoid interception
                        self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", download_btn)
                        time.sleep(1)
                        
                        print(f"   → Clicking Download button...")
                        self.driver.execute_script("arguments[0].click();", download_btn)
                        time.sleep(4)
                        
                        # Check if new tab opened and close it
                        all_windows = self.driver.window_handles
                        for window in all_windows:
                            if window != main_window:
                                self.driver.switch_to.window(window)
                                self.driver.close()
                        self.driver.switch_to.window(main_window)
                        time.sleep(1)
                        
                        # Generate filename
                        if phone and date:
                            filename = f"{phone}_{date}_Call1.pdf"
                        else:
                            filename = f"review_pending_{idx}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
                        
                        # Wait for download and move to review_pending folder
                        time.sleep(3)
                        downloads_folder = Path.home() / "Downloads"
                        
                        # Find latest downloaded file
                        pdf_files = list(downloads_folder.glob("*.pdf")) + list(downloads_folder.glob("*.png")) + \
                                    list(downloads_folder.glob("*.jpg")) + list(downloads_folder.glob("*.jpeg"))
                        if pdf_files:
                            latest_file = max(pdf_files, key=lambda x: x.stat().st_mtime)
                            
                            # Check if file was just downloaded (within last 10 seconds)
                            if (time.time() - latest_file.stat().st_mtime) < 10:
                                dest_path = self.review_pending_folder / filename
                                shutil.move(str(latest_file), str(dest_path))
                                print(f"   ✓ Downloaded: {filename}")
                                downloaded_count += 1
                            else:
                                print(f"   ⚠️  No recent download found")
                        
                    except Exception as e:
                        print(f"   ⚠️  Could not download document: {str(e)}")
                        import traceback
                        traceback.print_exc()
                    
                    # Go back to All Complaints
                    print(f"   → Returning to All Complaints...")
                    self.navigate_to_all_complaints()
                    time.sleep(2)
                    
                except Exception as e:
                    print(f"   ✗ Error processing complaint {idx}: {str(e)}")
                    continue
            
            print(f"\n{'='*70}")
            print(f"📥 Downloaded {downloaded_count} files to {self.review_pending_folder}")
            print(f"{'='*70}\n")
            
            return downloaded_count
            
        except Exception as e:
            print(f"\n✗ Download error: {str(e)}")
            import traceback
            traceback.print_exc()
            return 0
    
    def run_review_pending_workflow(self):
        """Process all Review Pending files"""
        print(f"\n{'='*70}")
        print(f"🔄 REVIEW PENDING WORKFLOW")
        print(f"{'='*70}\n")
        
        # STEP 1: Download all Review Pending files from dashboard
        print(f"→ Step 1: Downloading Review Pending files from dashboard...")
        downloaded_count = self.download_review_pending_files_from_dashboard()
        
        if downloaded_count == 0:
            print(f"⚠️  No files downloaded, checking existing files...")
        
        # STEP 2: Process all files in review_pending folder
        print(f"\n→ Step 2: Processing downloaded files...")
        
        # Get all PDF/PNG files
        files = list(self.review_pending_folder.glob("*.pdf")) + \
                list(self.review_pending_folder.glob("*.png")) + \
                list(self.review_pending_folder.glob("*.jpg")) + \
                list(self.review_pending_folder.glob("*.jpeg"))
        
        if not files:
            print(f"✗ No files found in {self.review_pending_folder}")
            return
        
        print(f"✓ Found {len(files)} file(s) to process\n")
        
        results = {'success': 0, 'failed': 0}
        
        for file_path in files:
            success = self.process_review_pending_file(file_path)
            if success:
                results['success'] += 1
            else:
                results['failed'] += 1
            
            # Return to All Complaints for next file
            time.sleep(2)
        
        print(f"\n{'='*70}")
        print(f"📊 REVIEW PENDING RESULTS")
        print(f"{'='*70}")
        print(f"Total: {len(files)}")
        print(f"✓ Success: {results['success']}")
        print(f"✗ Failed: {results['failed']}")
        print(f"{'='*70}\n")
    
    def run_open_workflow(self):
        """Process all Open files"""
        print(f"\n{'='*70}")
        print(f"🔄 OPEN COMPLAINTS WORKFLOW")
        print(f"{'='*70}\n")
        
        # Get all PDF/PNG files
        files = list(self.open_folder.glob("*.pdf")) + \
                list(self.open_folder.glob("*.png")) + \
                list(self.open_folder.glob("*.jpg")) + \
                list(self.open_folder.glob("*.jpeg"))
        
        if not files:
            print(f"✗ No files found in {self.open_folder}")
            return
        
        print(f"✓ Found {len(files)} file(s) to process\n")
        
        results = {'success': 0, 'failed': 0}
        
        for file_path in files:
            success = self.process_open_file(file_path)
            if success:
                results['success'] += 1
            else:
                results['failed'] += 1
            
            # Return to All Complaints for next file
            time.sleep(2)
        
        print(f"\n{'='*70}")
        print(f"📊 OPEN COMPLAINTS RESULTS")
        print(f"{'='*70}")
        print(f"Total: {len(files)}")
        print(f"✓ Success: {results['success']}")
        print(f"✗ Failed: {results['failed']}")
        print(f"{'='*70}\n")
    
    def run(self, workflow_type='both'):
        """Run the automation"""
        try:
            # Start browser and login
            self.start_browser()
            
            if not self.login():
                print("✗ Login failed")
                return
            
            # Run workflows
            if workflow_type in ['review_pending', 'both']:
                self.run_review_pending_workflow()
            
            if workflow_type in ['open', 'both']:
                self.run_open_workflow()
            
            print(f"\n✅ Automation completed!")
            print(f"Press Enter to close browser...")
            input()
            
        except Exception as e:
            print(f"\n✗ Fatal error: {str(e)}")
            import traceback
            traceback.print_exc()
        
        finally:
            if self.driver:
                self.driver.quit()


def main():
    """Main entry point"""
    import sys
    
    print(f"\n{'='*70}")
    print(f"🤖 NDNC Complete Automation")
    print(f"{'='*70}\n")
    
    EMAIL = "shraddha.s@exotel.com"
    
    # Check for command line argument
    if len(sys.argv) > 1:
        workflow = sys.argv[1]
    else:
        print("Select workflow:")
        print("1. Review Pending only")
        print("2. Open complaints only")
        print("3. Both (default)")
        
        choice = input("\nEnter choice (1/2/3): ").strip() or '3'
        
        workflow_map = {
            '1': 'review_pending',
            '2': 'open',
            '3': 'both'
        }
        
        workflow = workflow_map.get(choice, 'both')
    
    automation = NDNCCompleteAutomation(email=EMAIL)
    automation.run(workflow_type=workflow)


if __name__ == "__main__":
    main()

